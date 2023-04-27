import math
import requests
import csv
import logging
import time
import openpyxl
from functions.hash import generate_random_hash
from exceptions import NetworkException
from functions.intercepted_requests import intercepted_eol_request
from functions.session import session
from type import *

NO_UNIV_SCORE = False  # 是否不查询分数线
NO_ENROLL_PLAN = False  # 是否不查询招生计划
NO_MAJOR_SCORE = False  # 是否不查询专业分数线
PAGE_RANGE = []  # 大学列表页数范围
ITEM_OFFSET = 0  # 起始位置偏移，或者说从起始页的第n + 1个大学开始查询，只能为非负值
YEAR_SINCE = 2020  # 数据起始年份
QUERY_INTERVAL = 10  # 每次查询的间隔，单位为秒，低于10的值可能导致IP暂时被封
PROVINCE = '北京'  # 大学所在的省份，可以参考下面的PROVIENCE_DICT填写
GENERATE_XLSX = True  # 是否生成xlsx文件

# region 但是这是碰都不能碰的Region


PROVIENCE_DICT = {  # 数字ID（行政区划代码）到省份名的映射，比较搞笑的是eol.cn的接口用到的ID有的是数字有的是字符串
    11: '北京',
    12: '天津',
    13: '河北',
    14: '山西',
    15: '内蒙古',
    21: '辽宁',
    22: '吉林',
    23: '黑龙江',
    31: '上海',
    32: '江苏',
    33: '浙江',
    34: '安徽',
    35: '福建',
    36: '江西',
    37: '山东',
    41: '河南',
    42: '湖北',
    43: '湖南',
    44: '广东',
    45: '广西',
    46: '海南',
    50: '重庆',
    51: '四川',
    52: '贵州',
    53: '云南',
    54: '西藏',
    61: '陕西',
    62: '甘肃',
    63: '青海',
    64: '宁夏',
    65: '新疆'
}


# 把上面的字典进行一个反的转
REV_PROVIENCE_DICT = {v: k for k, v in PROVIENCE_DICT.items()}


HASH = generate_random_hash()  # 本次运行的HASH，用于标识CSV批次

logging.basicConfig(
    level=logging.INFO,
    datefmt="%m/%d %H:%M:%S",
    format='[%(asctime)s][%(levelname)s] %(message)s'
)


def load_dictionary() -> dict[str, str]:
    """加载用于渲染表格的字典"""
    try:
        res = session.get(
            'https://static-data.gaokao.cn/www/2.0/config/dicprovince/dic.json')
    except requests.exceptions.RequestException:
        raise NetworkException('加载字典时发生网络错误')
    return res.json()['data']


def get_univ_list(prov: str, rev_prov_dict: dict[str, int] = REV_PROVIENCE_DICT) -> list[Univ]:
    """Get the list of universities in a province."""
    prov_id: int = rev_prov_dict[prov]
    try:
        preflight_res = intercepted_eol_request(
            {
                'province_id': prov_id,
                'uri': 'apidata/api/gk/school/lists',
                'page': 1,
                'size': 1,
                'request_type': 1
            }
        )
    except requests.exceptions.RequestException:
        raise NetworkException('网络错误')

    item_count: int = preflight_res['numFound']
    page_count = math.ceil(item_count / 20)
    start_page = 1
    if len(PAGE_RANGE) == 2:
        page_count = min(PAGE_RANGE[1], math.ceil(item_count / 20))
        start_page = max(PAGE_RANGE[0], 1)

    univ_list = []

    for page in range(start_page, page_count + 1):
        try:
            res = intercepted_eol_request(
                {
                    'province_id': int(prov_id),
                    'uri': 'apidata/api/gk/school/lists',
                    'size': 20,
                    'page': page,
                    'request_type': 1
                }
            )
            time.sleep(QUERY_INTERVAL)

        except requests.exceptions.RequestException:
            raise NetworkException('网络错误')

        univ_list += res['item']

    return univ_list[ITEM_OFFSET:]


def get_minium_score_of_univ(univ: Univ, dictionary: dict[str, str], prov_dict: dict[int, str] = PROVIENCE_DICT) -> list[MiniumScoreForUnivs]:
    """获取高校各省各年份分数线"""
    school_id = univ['school_id']
    try:
        res = session.get(
            f'https://static-data.gaokao.cn/www/2.0/school/{school_id}/dic/provincescore.json'
        )
        if res.status_code == 404:
            # 某些学校，如军校，不公开招生，没有元数据可以爬
            logging.info('该校无信息，已跳过')
            return []
    except requests.exceptions.RequestException:
        raise NetworkException('网络错误')
    metadata: MetaMiniumScoreForUnivs = res.json()['data']

    ret_list: list[MiniumScoreForUnivs] = []

    for prov_id in metadata['newsdata']['province']:
        year_list = metadata['newsdata']['year'][str(prov_id)]
        year_list = filter(lambda x: x >= YEAR_SINCE, year_list)
        for year in year_list:
            for major_id in metadata['newsdata']['type'][f'{prov_id}_{year}']:
                try:
                    res = session.get(
                        f'https://static-data.gaokao.cn/www/2.0/schoolprovinceindex/{year}/{school_id}/{prov_id}/{major_id}/1.json')
                    if res.status_code != 200:
                        raise NetworkException('网络错误')
                except requests.exceptions.RequestException:
                    raise NetworkException('网络错误')
                data = res.json()['data']
                for item in data['item']:
                    new_row: MiniumScoreForUnivs = {
                        'code': univ['code_enroll'][:5],
                        'name': univ['name'],
                        'located_province': univ['province_name'],
                        'target_province': prov_dict[prov_id],
                        'major': dictionary[str(major_id)],
                        'year': year,
                        'enroll_level': item['local_batch_name'],
                        'enroll_type': item['zslx_name'],
                        'minium_score': item["min"],
                        'minium_rank': item["min_section"],
                        'prov_minium_score': item['proscore'],
                        'major_group': item['sg_name'] or None,
                        'subject_requirements': item['sg_info'] or None
                    }
                    ret_list.append(new_row)

    return ret_list


def get_enroll_plan_of_majors(univ: Univ, dictionary: dict[str, str], prov_dict: dict[int, str] = PROVIENCE_DICT) -> list[EnrollPlan]:
    """获取高校招生计划"""
    school_id = univ['school_id']
    try:
        res = session.get(
            f'https://static-data.gaokao.cn/www/2.0/school/{school_id}/dic/specialplan.json'
        )
        if res.status_code == 404:
            logging.info('该校无信息，已跳过')
            return []
    except requests.exceptions.RequestException:
        raise NetworkException('网络错误')
    metadata: MetaEnrollPlan = res.json()['data']

    ret_list: list[EnrollPlan] = []

    for prov_id in metadata['newsdata']['province']:
        year_list = metadata['newsdata']['year'][str(prov_id)]
        year_list = filter(lambda x: x >= YEAR_SINCE, year_list)
        for year in year_list:
            for major_id in metadata['newsdata']['type'][f'{prov_id}_{year}']:
                for batch_id in metadata['newsdata']['batch'][f'{prov_id}_{year}_{major_id}']:
                    try:
                        res = intercepted_eol_request(
                            {
                                'local_batch_id': batch_id,
                                'local_province_id': prov_id,
                                'local_type_id': str(major_id),
                                'page': 1,
                                'school_id': school_id,
                                'size': 30,
                                'uri': 'apidata/api/gkv3/plan/school',
                                'year': year
                            }
                        )
                    except requests.exceptions.RequestException:
                        raise NetworkException('网络错误')

                    total_items = res['numFound']
                    total_pages = math.ceil(total_items / 30)

                    for page in range(1, total_pages + 1):
                        try:
                            time.sleep(QUERY_INTERVAL)
                            res = intercepted_eol_request(
                                {
                                    'local_batch_id': batch_id,
                                    'local_province_id': prov_id,
                                    'local_type_id': str(major_id),
                                    'page': page,
                                    'school_id': school_id,
                                    'size': 30,
                                    'uri': 'apidata/api/gkv3/plan/school',
                                    'year': year
                                }
                            )
                        except requests.exceptions.RequestException:
                            raise NetworkException('网络错误')

                        data = res

                        for item in data['item']:
                            new_row: EnrollPlan = {
                                'code': univ['code_enroll'][:5],
                                'name': univ['name'],
                                'located_province': univ['province_name'],
                                'target_province': prov_dict[prov_id],
                                'year': year,
                                'major': dictionary[str(major_id)],
                                'enroll_level': data['item'][0]['local_batch_name'],
                                'major_name': item['spname'],
                                'planned_number': item['num'],
                                'duration': item['length'],
                                'tuition': item['tuition'],
                                'major_group': item['sg_name'] or None,
                                'subject_requirements': item['sg_info'] or None,
                            }
                            ret_list.append(new_row)

    return ret_list


def get_minium_score_of_majors(univ: Univ, dictionary: dict[str, str], prov_dict: dict[int, str] = PROVIENCE_DICT) -> list[MiniumScoreForMajors]:
    """获取高校专业分数线"""
    school_id = univ['school_id']
    try:
        res = session.get(
            f'https://static-data.gaokao.cn/www/2.0/school/{school_id}/dic/specialscore.json'
        )
        if res.status_code == 404:
            logging.info('该校无信息，已跳过')
            return []
    except requests.exceptions.RequestException:
        raise NetworkException('网络错误')
    metadata: MetaMiniumScoreForMajors = res.json()['data']

    ret_list: list[MiniumScoreForMajors] = []

    for prov_id in metadata['newsdata']['province']:
        year_list = metadata['newsdata']['year'][str(prov_id)]
        year_list = filter(lambda x: x >= YEAR_SINCE, year_list)
        for year in year_list:
            for major_id in metadata['newsdata']['type'][f'{prov_id}_{year}']:
                for batch_id in metadata['newsdata']['batch'][f'{prov_id}_{year}_{major_id}']:
                    try:
                        res = intercepted_eol_request(
                            {
                                'local_batch_id': batch_id,
                                'local_province_id': prov_id,
                                'local_type_id': str(major_id),
                                'page': 1,
                                'school_id': school_id,
                                'size': 30,
                                'uri': 'apidata/api/gk/score/special',
                                'year': year
                            }
                        )
                    except requests.exceptions.RequestException:
                        raise NetworkException('网络错误')

                    total_items = res['numFound']
                    total_pages = math.ceil(total_items / 30)

                    for page in range(1, total_pages + 1):
                        try:
                            time.sleep(QUERY_INTERVAL)
                            res = intercepted_eol_request(
                                {
                                    'local_batch_id': batch_id,
                                    'local_province_id': prov_id,
                                    'local_type_id': str(major_id),
                                    'page': page,
                                    'school_id': school_id,
                                    'size': 30,
                                    'uri': 'apidata/api/gk/score/special',
                                    'year': year
                                }
                            )
                        except requests.exceptions.RequestException:
                            raise NetworkException('网络错误')

                        data = res

                        for item in data['item']:
                            new_row: MiniumScoreForMajors = {
                                'code': univ['code_enroll'][:5],
                                'name': univ['name'],
                                'located_province': univ['province_name'],
                                'target_province': prov_dict[prov_id],
                                'year': year,
                                'major': dictionary[str(major_id)],
                                'major_name': item['spname'],
                                'enroll_level': data['item'][0]['local_batch_name'],
                                'avg_score': item['average'],
                                'minium_score': item["min"],
                                'minium_rank': item["min_section"],
                                'major_group': item['sg_name'] or None,
                                'subject_requirements': item['sg_info'] or None
                            }
                            ret_list.append(new_row)
    return ret_list


def main():
    try:
        logging.info('开始获取大学列表')
        univ_list: list[Univ] = get_univ_list(PROVINCE)
    except NetworkException:
        logging.fatal('获取大学列表时发生网络异常')
        exit(1)

    try:
        logging.info('开始获取专业字典')
        dictionary = load_dictionary()
    except NetworkException:
        logging.fatal('获取专业字典时发生网络异常')
        exit(1)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore

    if not NO_UNIV_SCORE:
        logging.info('开始获取高校各省分数线')
        if GENERATE_XLSX:
            wb.create_sheet('学校分数线')
            wb['学校分数线'].append([
                '学校代码',
                '学校名称全称',
                '所在省份',
                '面向省份',
                '科类',
                '年份',
                '录取批次',
                '招生类型',
                '最低分',
                '最低位次',
                '省控线',
                '专业组',
                '选科要求'
            ])
        with open(f'min_score_{HASH}.csv', 'w', encoding='utf-8', newline='') as csvfile:
            csvwriter = csv.DictWriter(
                csvfile, fieldnames=MiniumScoreForUnivs.__annotations__.keys()
            )
            csvwriter.writeheader()
            for univ in univ_list:
                try:
                    logging.info(f'正在获取{univ["name"]}分数线')
                    score_list = get_minium_score_of_univ(univ, dictionary)
                except NetworkException:
                    logging.fatal(f'获取{univ["name"]}的分数线时发生网络异常')
                    exit(1)
                for score in score_list:
                    csvwriter.writerow(score)
                    if GENERATE_XLSX:
                        wb['学校分数线'].append([v for v in score.values()])
                if GENERATE_XLSX:
                    wb.save(f'data_{HASH}.xlsx')
                logging.info(f'成功获取{univ["name"]}分数线信息')
        logging.info('已获取全部高校分数线信息')

    if not NO_ENROLL_PLAN:
        logging.info('开始获取高校各专业招生计划')
        if GENERATE_XLSX:
            wb.create_sheet('各专业招生计划')
            wb['各专业招生计划'].append([
                '学校代码',
                '学校名称全称',
                '所在省份',
                '面向省份',
                '年份',
                '科类',
                '招生批次',
                '招生专业名称',
                '计划招生',
                '学制',
                '学费',
                '专业组',
                '选科要求'
            ])
        with open(f'enroll_plan_{HASH}.csv', 'w', encoding='utf-8', newline='') as csvfile:
            csvwriter = csv.DictWriter(
                csvfile, fieldnames=EnrollPlan.__annotations__.keys()
            )
            csvwriter.writeheader()
            for univ in univ_list:
                try:
                    logging.info(f'正在获取{univ["name"]}招生计划')
                    plan_list = get_enroll_plan_of_majors(univ, dictionary)
                except NetworkException:
                    logging.fatal(f'获取{univ["name"]}的招生计划时发生网络异常')
                    exit(1)
                for plan in plan_list:
                    csvwriter.writerow(plan)
                    if GENERATE_XLSX:
                        wb['各专业招生计划'].append([v for v in plan.values()])
                if GENERATE_XLSX:
                    wb.save(f'data_{HASH}.xlsx')
                logging.info(f'成功获取{univ["name"]}招生计划信息')
        logging.info('已获取全部高校招生计划信息')

    if not NO_MAJOR_SCORE:
        logging.info('开始获取高校各专业分数线')
        if GENERATE_XLSX:
            wb.create_sheet('分专业录取分数线')
            wb['分专业录取分数线'].append([
                '学校代码',
                '学校名称全称',
                '所在省份',
                '面向省份',
                '年份',
                '科类',
                '录取专业名称',
                '录取批次',
                '平均分',
                '最低分',
                '最低位次',
                '专业组',
                '选科要求'
            ])
        with open(f'major_score_{HASH}.csv', 'w', encoding='utf-8', newline='') as csvfile:
            csvwriter = csv.DictWriter(
                csvfile, fieldnames=MiniumScoreForMajors.__annotations__.keys()
            )
            csvwriter.writeheader()
            for univ in univ_list:
                try:
                    logging.info(f'正在获取{univ["name"]}各专业分数线')
                    major_score = get_minium_score_of_majors(univ, dictionary)
                except NetworkException:
                    logging.fatal(f'获取{univ["name"]}的各专业分数线时发生网络异常')
                    exit(1)
                for score in major_score:
                    csvwriter.writerow(score)
                    wb['分专业录取分数线'].append([v for v in score.values()])
                if GENERATE_XLSX:
                    wb.save(f'data_{HASH}.xlsx')
                logging.info(f'成功获取{univ["name"]}各专业分数线信息')
        logging.info('已获取全部高校各专业分数线信息')

    logging.info('已成功爬取所有数据')


if __name__ == '__main__':
    main()

# endregion
