import click
import openpyxl
import csv as _csv
from type import MiniumScoreForUnivs, EnrollPlan, MiniumScoreForMajors
from typing import TypedDict, cast
from main import HASH


class Form(TypedDict):
    """完整表格"""
    province: list[MiniumScoreForUnivs]
    """省分数线"""
    major: list[MiniumScoreForMajors]
    """专业分数线"""
    enroll: list[EnrollPlan]
    """招生计划"""


HEADERS = (
    'code,name,located_province,target_province,major,year,enroll_level,enroll_type,minium_score_and_rank,prov_minium_score,major_group,major_requirements',  # 省分数线
    'code,name,located_province,target_province,year,major,enroll_level,major_name,planned_number,duration,tuition,major_requirements',  # 招生计划
    'code,name,located_province,target_province,year,major,major_name,enroll_level,avg_score,minium_score_and_rank,major_requirements'  # 专业分数线
)

READABLE_HEADERS = {
    'province': [
        '学校代码',
        '学校名称全称',
        '所在省份',
        '面向省份',
        '科类',
        '年份',
        '录取批次',
        '招生类型',
        '最低分/最低位次',
        '省控线',
        '专业组',
        '选科要求'
    ],
    'enroll': [
        '学校代码',
        '学校名称全称',
        '所在省份',
        '面向省份',
        '年份',
        '科类',
        '招生批次',
        '招生专业名称',
        '计划招生',
        '学制',
        '学费',
        '选科要求'
    ],
    'major': [
        '学校代码',
        '学校名称全称',
        '所在省份',
        '面向省份',
        '年份',
        '科类',
        '录取专业名称',
        '录取批次',
        '平均分',
        '最低分/最低位次',
        '选科要求'
    ]
}

NAME_DICT = {
    'province': '学校分数线',
    'enroll': '各专业招生计划',
    'major': '分专业录取分数线'
}


@click.command('merge', help='合并多个CSV或XLSX文件，也可用于转换表格格式')
@click.option(
    '--csv', '-c',
    type=click.Path(
        exists=True, file_okay=True,
        readable=True, resolve_path=True
    ),
    multiple=True,
    help='CSV文件路径，此选项可重复使用'
)
@click.option(
    '--xlsx', '-x',
    type=click.Path(
        exists=True, file_okay=True,
        readable=True, resolve_path=True
    ),
    multiple=True,
    help='XLSX文件路径，此选项可重复使用'
)
@click.option(
    '--type', '-t',
    type=click.Choice(['csv', 'xlsx']),
    help='输出文件类型，默认为csv，输出的文件将位于当前工作目录下',
    default='csv'
)
@click.option(
    '--remove-empty-lines/--no-remove-empty-lines',
    default=True,
    help='是否使生成文件中不包含来自CSV源文件的空行，默认开启'
)
def merge(csv: list[str], xlsx: list[str], type: str, remove_empty_lines: bool):
    form: Form = {
        'province': [],
        'enroll': [],
        'major': []
    }

    if not csv and not xlsx:
        ctx = click.get_current_context()
        ctx.fail('请至少指定一个CSV或XLSX文件')

    for file in csv:
        with open(file, 'r', encoding='utf-8') as f:
            """你用cast强制转类型的样子真的很狼狈"""
            reader = _csv.DictReader(f)  # type: ignore
            header = ','.join(reader.fieldnames)  # type: ignore
            if header == HEADERS[0]:
                """省分数线"""
                form['province'].extend(list(reader))  # type: ignore
            elif header == HEADERS[1]:
                """招生计划"""
                form['enroll'].extend(list(reader))  # type: ignore
            elif header == HEADERS[2]:
                """专业分数线"""
                form['major'].extend(list(reader))  # type: ignore

    for file in xlsx:
        wb = openpyxl.load_workbook(file)
        if '学校分数线' in wb:
            for row in wb['学校分数线'].rows:
                if row[0].value == '学校代码':
                    continue
                new_row = cast(MiniumScoreForUnivs,
                               {k: v for k, v in zip(
                                   MiniumScoreForUnivs.__annotations__.keys(),
                                   (v.value for v in row)
                               )})
                form['province'].append(new_row)
        if '各专业招生计划' in wb:
            for row in wb['各专业招生计划'].rows:
                if row[0].value == '学校代码':
                    continue
                new_row = cast(EnrollPlan,
                               {k: v for k, v in zip(
                                   EnrollPlan.__annotations__.keys(),
                                   (v.value for v in row)
                               )})
                form['enroll'].append(new_row)
        if '分专业录取分数线' in wb:
            for row in wb['分专业录取分数线'].rows:
                if row[0].value == '学校代码':
                    continue
                new_row = cast(MiniumScoreForMajors,
                               {k: v for k, v in zip(
                                   MiniumScoreForMajors.__annotations__.keys(),
                                   (v.value for v in row)
                               )})
                form['major'].append(new_row)

    if remove_empty_lines:
        for k, v in form.items():
            v = cast(list, v)  # 这个语法真的好丑
            form[k] = list(filter(lambda x: x, v))

    if type == 'csv':
        for k, v in form.items():
            if v:
                v = cast(list, v)
                with open(f'{k}_output_{HASH}.csv', 'w', encoding='utf-8', newline='') as f:
                    writer = _csv.DictWriter(f, fieldnames=v[0].keys())
                    writer.writeheader()
                    writer.writerows(v)
    elif type == 'xlsx':
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for k, v in form.items():
            if v:
                v = cast(list, v)
                ws = wb.create_sheet(NAME_DICT[k])
                ws.append(READABLE_HEADERS[k])
                for row in v:
                    ws.append(list(row.values()))
        wb.save(f'output_{HASH}.xlsx')


if __name__ == '__main__':
    merge()
